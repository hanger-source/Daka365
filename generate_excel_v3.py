import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import calendar
from datetime import date

# 定义样式
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ACHIEVEMENT_ROW_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
BORDER_STYLE = Border(left=Side(style='thin', color='C0C0C0'),
                      right=Side(style='thin', color='C0C0C0'),
                      top=Side(style='thin', color='C0C0C0'),
                      bottom=Side(style='thin', color='C0C0C0'))

def generate_linked_365_excel(filename="365天打卡模板_v3_联动版.xlsx", year=2026, max_items=50):
    workbook = openpyxl.Workbook()
    
    # --- 1. 创建配置页 (Settings) ---
    config_sheet = workbook.active
    config_sheet.title = "配置页"
    
    config_headers = ["序号", "类别", "事项", "目标"]
    config_sheet.row_dimensions[1].height = 30
    for i, h in enumerate(config_headers, 1):
        cell = config_sheet.cell(row=1, column=i)
        cell.value = h
        cell.font = Font(bold=True, size=14) # 加大表头字体
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_STYLE
    
    # 初始数据
    initial_items = [
        [1, "生活", "早睡早起", 21],
        [2, "生活", "跑步", 21],
        [3, "生活", "每天护肤", 31],
        [4, "生活", "吃早餐", 31],
        [5, "学习", "睡前阅读半小时", 10],
        [6, "工作", "剪视频", 14]
    ]
    
    # 为配置页所有可能的事项行（1-50行）预设格式
    for r_idx in range(2, max_items + 2):
        config_sheet.row_dimensions[r_idx].height = 25 # 设置行高
        for c_idx in range(1, 5):
            cell = config_sheet.cell(row=r_idx, column=c_idx)
            # 填入初始数据
            if r_idx - 2 < len(initial_items):
                cell.value = initial_items[r_idx - 2][c_idx - 1]
            
            # 核心修复：预设所有格子的样式，保证手动添加也居中
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_STYLE

    config_sheet.column_dimensions['A'].width = 10
    config_sheet.column_dimensions['B'].width = 20
    config_sheet.column_dimensions['C'].width = 35
    config_sheet.column_dimensions['D'].width = 15

    # --- 2. 创建 12 个月的打卡页 ---
    for month_num in range(1, 13):
        sheet_title = f"{year}年{month_num}月打卡"
        sheet = workbook.create_sheet(title=sheet_title)

        num_days = calendar.monthrange(year, month_num)[1]
        ach_col_idx = num_days + 5  
        mom_col_idx = num_days + 6  
        ach_col_letter = get_column_letter(ach_col_idx)
        mom_col_letter = get_column_letter(mom_col_idx)
        col_e_start, col_e_end = get_column_letter(5), get_column_letter(num_days + 4)

        # 设置表头
        sheet.row_dimensions[1].height = 35
        sheet.row_dimensions[2].height = 35

        # 固定列表头
        headers = [("A1", "序号"), ("B1", "类别"), ("C1", "事项"), ("D1", "目标")]
        for cell_ref, label in headers:
            sheet.merge_cells(f"{cell_ref}:{cell_ref[0]}2")
            cell = sheet[cell_ref]
            cell.value = label
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = HEADER_FILL
            cell.border = BORDER_STYLE
            sheet[f"{cell_ref[0]}2"].border = BORDER_STYLE

        # 日期表头
        for d in range(1, num_days + 1):
            col_letter = get_column_letter(d + 4)
            day_of_week = date(year, month_num, d).weekday()
            weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            sheet[f"{col_letter}1"].value = weekdays[day_of_week]
            sheet[f"{col_letter}2"].value = d
            for r in [1, 2]:
                cell = sheet[f"{col_letter}{r}"]
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = HEADER_FILL
                cell.border = BORDER_STYLE

        # 指标表头
        for idx, label in [(ach_col_idx, "达成率"), (mom_col_idx, "环比增长率")]:
            col_letter = get_column_letter(idx)
            sheet.merge_cells(f"{col_letter}1:{col_letter}2")
            cell = sheet[f"{col_letter}1"]
            cell.value = label
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = HEADER_FILL
            cell.border = BORDER_STYLE
            sheet[f"{col_letter}2"].border = BORDER_STYLE
            sheet.column_dimensions[col_letter].width = 18

        # --- 事项内容填充 (引用配置页) ---
        start_row = 3
        end_row = max_items + 2
        
        for i in range(max_items):
            row = i + 3
            config_row = i + 2
            sheet.row_dimensions[row].height = 30
            
            # 引用公式：锁定配置页的事项名(C列)作为判断基准
            sheet[f"A{row}"].value = f"=IF(配置页!$C${config_row}<>\"\", 配置页!A{config_row}, \"\")"
            sheet[f"B{row}"].value = f"=IF(配置页!$C${config_row}<>\"\", 配置页!B{config_row}, \"\")"
            sheet[f"C{row}"].value = f"=IF(配置页!$C${config_row}<>\"\", 配置页!C{config_row}, \"\")"
            sheet[f"D{row}"].value = f"=IF(配置页!$C${config_row}<>\"\", 配置页!D{config_row}, \"\")"

            for col in ["A", "B", "C", "D"]:
                cell = sheet[f"{col}{row}"]
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER_STYLE

            # 打卡格子
            for d_col in range(5, num_days + 5):
                cell = sheet.cell(row=row, column=d_col)
                cell.border = BORDER_STYLE
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 达成率
            ach_c = sheet[f"{ach_col_letter}{row}"]
            ach_c.value = f'=IF(C{row}<>"", IFERROR(COUNTIF({col_e_start}{row}:{col_e_end}{row},"<>" )/D{row}, 0), "")'
            ach_c.number_format = '0.00%'
            ach_c.font = Font(size=12)
            ach_c.alignment, ach_c.border = Alignment(horizontal='center', vertical='center'), BORDER_STYLE

            # 环比
            mom_c = sheet[f"{mom_col_letter}{row}"]
            if month_num > 1:
                prev = f"{year}年{month_num-1}月打卡"
                prev_ach_col = get_column_letter(calendar.monthrange(year, month_num-1)[1] + 5)
                mom_c.value = f'=IF(C{row}<>"", IFERROR(({ach_col_letter}{row}-\'{prev}\'!{prev_ach_col}{row})/\'{prev}\'!{prev_ach_col}{row}, 0), "")'
            mom_c.number_format, mom_c.alignment, mom_c.border = '0.00%', Alignment(horizontal='center', vertical='center'), BORDER_STYLE
            mom_c.font = Font(size=12)

        # --- 每日备注 & 汇总区域 ---
        note_row = end_row + 1
        sheet.row_dimensions[note_row].height = 35
        sheet.merge_cells(f"A{note_row}:D{note_row}")
        sheet[f"A{note_row}"] = "每日备注"
        sheet[f"A{note_row}"].font, sheet[f"A{note_row}"].fill = Font(bold=True, size=14), HEADER_FILL
        sheet[f"A{note_row}"].alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, mom_col_idx + 1):
            cell = sheet.cell(row=note_row, column=c)
            cell.border = BORDER_STYLE
            if c > 4: cell.fill = ACHIEVEMENT_ROW_FILL

        summary_start = note_row + 2
        sheet.merge_cells(start_row=summary_start, end_row=summary_start, start_column=1, end_column=4)
        title_c = sheet[f"A{summary_start}"]
        title_c.value, title_c.font, title_c.fill = "月度数据汇总", Font(bold=True, size=16), HEADER_FILL
        title_c.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, 5): sheet.cell(row=summary_start, column=c).border = BORDER_STYLE

        item_header_row = summary_start + 1
        sheet.merge_cells(f"A{item_header_row}:B{item_header_row}")
        sheet[f"A{item_header_row}"], sheet[f"C{item_header_row}"], sheet[f"D{item_header_row}"] = "事项明细", "打卡天数", "达成率"
        for c in range(1, 5):
            cell = sheet.cell(row=item_header_row, column=c)
            cell.font, cell.fill, cell.border = Font(bold=True, size=13), ACHIEVEMENT_ROW_FILL, BORDER_STYLE
            cell.alignment = Alignment(horizontal='center', vertical='center')

        curr_r = item_header_row + 1
        for i in range(max_items):
            row_in_data = i + 3
            sheet.merge_cells(f"A{curr_r}:B{curr_r}")
            sheet[f"A{curr_r}"] = f"=IF(C{row_in_data}<>\"\", C{row_in_data}, \"\")"
            sheet[f"C{curr_r}"] = f'=IF(C{row_in_data}<>"", COUNTIF({col_e_start}{row_in_data}:{col_e_end}{row_in_data},"<>" ), "")'
            sheet[f"D{curr_r}"] = f'=IF(C{row_in_data}<>"", {ach_col_letter}{row_in_data}, "")'
            sheet[f"D{curr_r}"].number_format = '0.00%'
            for c in range(1, 5):
                cell = sheet.cell(row=curr_r, column=c)
                cell.border = BORDER_STYLE
                cell.alignment = Alignment(horizontal='center', vertical='center')
            curr_r += 1

        total_header_row = curr_r
        sheet.merge_cells(f"A{total_header_row}:B{total_header_row}")
        sheet[f"A{total_header_row}"], sheet[f"C{total_header_row}"], sheet[f"D{total_header_row}"] = "统计指标", "数值", "环比增长"
        for c in range(1, 5):
            cell = sheet.cell(row=total_header_row, column=c)
            cell.font, cell.fill, cell.border = Font(bold=True, size=13), ACHIEVEMENT_ROW_FILL, BORDER_STYLE
            cell.alignment = Alignment(horizontal='center', vertical='center')

        labels = ["事项总数:", "已完成事项(达成率>0):", "月度平均达成率:", "总打卡天数:", "总有效打卡比例:"]
        for i, label in enumerate(labels):
            r = total_header_row + 1 + i
            sheet.merge_cells(f"A{r}:B{r}")
            sheet[f"A{r}"] = label
            val_c, sum_mom_c = sheet[f"C{r}"], sheet[f"D{r}"]
            val_c.alignment = sum_mom_c.alignment = Alignment(horizontal='center', vertical='center')
            val_c.border = sum_mom_c.border = BORDER_STYLE
            sheet[f"A{r}"].border = BORDER_STYLE

            if i == 0: val_c.value = f'=COUNTA(C{start_row}:C{end_row})'
            elif i == 1: val_c.value = f'=COUNTIF({ach_col_letter}{start_row}:{ach_col_letter}{end_row},">0")'
            elif i == 2: val_c.value, val_c.number_format = f'=IFERROR(AVERAGE({ach_col_letter}{start_row}:{ach_col_letter}{end_row}),0)', '0.00%'
            elif i == 3: val_c.value = f'=COUNTIF({col_e_start}{start_row}:{col_e_end}{end_row},"<>" )'
            elif i == 4: val_c.value, val_c.number_format = f'=IFERROR(C{r-1}/(C{r-4}*{num_days}),0)', '0.00%'

            if month_num > 1:
                prev_s = f"{year}年{month_num-1}月打卡"
                sum_mom_c.value = f'=IFERROR((C{r}-\'{prev_s}\'!C{r})/\'{prev_s}\'!C{r},0)'
                sum_mom_c.number_format = '0.00%'

        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 12
        for d in range(5, num_days + 5): sheet.column_dimensions[get_column_letter(d)].width = 8

    workbook.save(filename)
    print(f"联动修复版模板 '{filename}' 已生成！\n《配置页》前 50 行已设好居中格式。")

if __name__ == "__main__":
    generate_linked_365_excel()
