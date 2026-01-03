import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
import calendar
from datetime import date, timedelta

# ==========================================
# 🎨 全局样式与色系配置
# ==========================================
class TrackerTheme:
    TEXT_COLOR = "000000"
    BORDER_COLOR = "94A3B8"
    DASHBOARD_COLOR = "D8E2DC"
    SUMMARY_LABEL_COLOR = "FFE5D9"
    HEADER_COLOR = "ECE4DB"
    HEADER_ALT_COLOR = "D6CCC2"
    ZEBRA_COLOR = "F8FAFC"
    REMARK_COLOR = "F1F5F9"
    SUCCESS_BG_COLOR = "D1FAE5"
    SUCCESS_TEXT_COLOR = "059669"
    ERROR_BG_COLOR = "FEE2E2"
    ERROR_TEXT_COLOR = "B91C1C"
    WEEKEND_FILL_COLOR = "F1F5F9"
    
    SCALE_RED = "FCA5A5"
    SCALE_YELLOW = "FDE68A"
    SCALE_GREEN = "86EFAC"
    SCALE_WHITE = "FFFFFF"

    @classmethod
    def get_border(cls, color=None):
        side = Side(style='thin', color=color or cls.BORDER_COLOR)
        return Border(left=side, right=side, top=side, bottom=side)

    @classmethod
    def get_fill(cls, color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    @classmethod
    def get_no_border(cls):
        return Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

class HabitTrackerGenerator:
    def __init__(self, filename="365天打卡模板_v5_正式版.xlsx", year=2026, max_items=50):
        self.filename = filename
        self.year = year
        self.max_items = max_items
        self.wb = openpyxl.Workbook()
        self.theme = TrackerTheme
        self.row_offset = 2
        self.col_offset = 2
        self.main_table_start = 14 + self.row_offset # 行 16
        self.remark_row = 11 + self.row_offset # 行 13
        
    def generate(self):
        self._setup_config_sheet()
        self._setup_annual_summary_sheet()
        self._setup_monthly_sheets()
        self._save()

    def _apply_common_settings(self, sheet):
        sheet.sheet_view.showGridLines = False
        # 🚨 已调回：设置工作表默认缩放比例为 100%
        sheet.sheet_view.zoomScale = 100
        sheet.protection.set_password('123456')
        sheet.protection.sheet = True
        sheet.protection.formatCells = False; sheet.protection.insertRows = False; sheet.protection.deleteRows = False; sheet.protection.sort = False; sheet.protection.autoFilter = False

    def _setup_config_sheet(self):
        ws = self.wb.active
        ws.title = "事项配置页"
        self._apply_common_settings(ws)
        
        headers = ["序号", "类别", "事项", "目标天数", "积极标志"]
        ws.row_dimensions[1].height = 40
        h_fill, h_border = self.theme.get_fill(self.theme.DASHBOARD_COLOR), self.theme.get_border()
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, size=14); cell.fill = h_fill; cell.border = h_border; cell.alignment = Alignment(horizontal='center', vertical='center')

        for r in range(2, self.max_items + 2):
            ws.row_dimensions[r].height = 35
            for c in range(1, 27): 
                cell = ws.cell(row=r, column=c); cell.protection = Protection(locked=False)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(c==3))
                if c == 1: cell.value = f'=IF(C{r}<>"", ROW()-1, "")'
                # 积极标志列（E列）字体调大
                if c == 5: cell.font = Font(size=20)

        ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 12; ws.column_dimensions['C'].width = 45; ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 15
        
        dv = DataValidation(type="custom", formula1=f'COUNTIF($C$2:$C${self.max_items+1}, C2)<=1', showErrorMessage=True, errorStyle="stop")
        dv.errorTitle, dv.error = "❌ 事项重复", "该事项已经存在！请勿重复添加。"
        ws.add_data_validation(dv); dv.add(f"C2:C{self.max_items+1}")

        full_row_range = f"A2:Z{self.max_items + 1}"; visible_data_range = f"A2:E{self.max_items + 1}"
        ws.conditional_formatting.add(full_row_range, FormulaRule(formula=[f'$C2=""'], font=Font(color="FFFFFF"), fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), border=self.theme.get_no_border(), stopIfTrue=True))
        ws.conditional_formatting.add(visible_data_range, FormulaRule(formula=['$C2<>""'], border=self.theme.get_border()))

        for r_idx, row_data in enumerate([["生活", "早睡早起", 21, "✅🔥"], ["生活", "跑步", 21, "🏃‍♂️💪"], ["学习", "睡前阅读", 10, "📖💡"]], 2):
            for c_idx, val in enumerate(row_data, 2): ws.cell(row=r_idx, column=c_idx, value=val)

    def _setup_annual_summary_sheet(self):
        ws = self.wb.create_sheet("📅 年度汇总看板", 1)
        self._apply_common_settings(ws)
        ws.freeze_panes = None 
        
        c_border, l_fill, d_fill, h_fill = self.theme.get_border(), self.theme.get_fill(self.theme.SUMMARY_LABEL_COLOR), self.theme.get_fill(self.theme.DASHBOARD_COLOR), self.theme.get_fill(self.theme.HEADER_COLOR)

        dash_y = 1 + self.row_offset
        ws.merge_cells(start_row=dash_y, start_column=1+self.col_offset, end_row=dash_y, end_column=76+self.col_offset) 
        title_cell = ws.cell(row=dash_y, column=1+self.col_offset, value="🏆 全年度打卡看板")
        title_cell.font = Font(bold=True, size=26); title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[dash_y].height = 60 
        
        # --- 全指标排版 (4 + 3 两行排列，防止超出屏幕) ---
        metric_cols = [1+self.col_offset + i*17 for i in range(4)] # 增大间距，每项占 10 列，间隔 7 列
        
        # 预先生成公式部分
        avg_rate_formula = ",".join([f"'{self.year}年{m}月打卡'!E7" for m in range(1, 13)])
        avg_pos_formula = ",".join([f"'{self.year}年{m}月打卡'!E8" for m in range(1, 13)])
        total_hits_formula = ",".join([f"SUM('{self.year}年{m}月打卡'!H{self.main_table_start}:H{self.main_table_start+self.max_items-1})" for m in range(1, 13)])
        active_days_formula = ",".join([f"'{self.year}年{m}月打卡'!E9" for m in range(1, 13)])

        metrics = [
            ("年度总事项", f'=COUNTIF(事项配置页!$C$2:$C${self.max_items+1}, "?*")', None),
            ("年度平均打卡率", f"=IFERROR(AVERAGE({avg_rate_formula}), 0)", '0.0%'),
            ("年度平均积极率", f"=IFERROR(AVERAGE({avg_pos_formula}), 0)", '0.0%'),
            ("累计打卡总次数", f"=SUM({total_hits_formula})", None),
            ("累计活跃总天数", f"=SUM({active_days_formula})", None),
            ("月均打卡次数", f"=IFERROR(SUM({total_hits_formula})/12, 0)", '0.0'),
            ("年度综合评分", f"=({get_column_letter(metric_cols[2])}{dash_y+3}*40) + ({get_column_letter(metric_cols[1])}{dash_y+3}*20) + ({get_column_letter(metric_cols[0])}{dash_y+7}/365*20) + (IFERROR({get_column_letter(metric_cols[2])}{dash_y+3}/{get_column_letter(metric_cols[1])}{dash_y+3}, 0)*20)", '0.0')
        ]

        for i, (label, formula, fmt) in enumerate(metrics):
            # 前 4 个在第一行 (dash_y+2)，后 3 个在第二行 (dash_y+6)
            row_off = 2 if i < 4 else 6
            curr_c = metric_cols[i % 4]
            
            ws.merge_cells(start_row=dash_y+row_off, start_column=curr_c, end_row=dash_y+row_off, end_column=curr_c+9)
            ws.cell(row=dash_y+row_off, column=curr_c, value=label).fill = l_fill; ws.cell(row=dash_y+row_off, column=curr_c).alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=dash_y+row_off+1, start_column=curr_c, end_row=dash_y+row_off+2, end_column=curr_c+9)
            val_cell = ws.cell(row=dash_y+row_off+1, column=curr_c, value=formula)
            val_cell.font = Font(bold=True, size=28); val_cell.alignment = Alignment(horizontal='center', vertical='center'); val_cell.border = c_border
            if fmt: val_cell.number_format = fmt

        # 画廊排版调整
        gallery_start_y = dash_y + 11 
        block_width, block_height = 34, 18 
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 3
        for c in range(1, 120):
            col_let = get_column_letter(c)
            if c <= self.col_offset: continue 
            adj_c = c - self.col_offset
            if (adj_c-1) % block_width == 0: ws.column_dimensions[col_let].width = 8 
            elif (adj_c-1) % block_width < 32: ws.column_dimensions[col_let].width = 4.5 # 进一步微调列宽
            else: ws.column_dimensions[col_let].width = 6 

        for i in range(self.max_items):
            col_idx, row_idx, cfg_r = (i % 2) * block_width + 1 + self.col_offset, (i // 2) * block_height + gallery_start_y, i + 2
            ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx + 31)
            title_cell = ws.cell(row=row_idx, column=col_idx, value=f'=IF(事项配置页!$C${cfg_r}<>"", "🔥 " & 事项配置页!$C${cfg_r}, "")')
            title_cell.font = Font(bold=True, size=16); title_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row_idx].height = 35

            block_range = f"{get_column_letter(col_idx)}{row_idx}:{get_column_letter(col_idx+31)}{row_idx+13}"
            ws.conditional_formatting.add(block_range, FormulaRule(formula=[f'事项配置页!$C${cfg_r}=""'], font=Font(color="FFFFFF"), fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), border=self.theme.get_no_border(), stopIfTrue=True))

            for d in range(1, 32):
                c = col_idx + d; cell = ws.cell(row=row_idx+1, column=c, value=d)
                cell.fill = h_fill; cell.border = c_border; cell.alignment = Alignment(horizontal='center', vertical='center')

            for m in range(1, 13):
                r = row_idx + 1 + m; ws.row_dimensions[r].height = 22.5 # 微调行高为 22.5，保持正方形比例
                ws.cell(row=r, column=col_idx, value=f"{m}月").fill = h_fill; ws.cell(row=r, column=col_idx).border = c_border; ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                m_name = f"{self.year}年{m}月打卡"; num_days = calendar.monthrange(self.year, m)[1]
                for d in range(1, 32):
                    curr_c = col_idx + d
                    if d <= num_days:
                        # 引用月度表中 M 列以后的每日打卡数据
                        formula = f'=IF(事项配置页!$C${cfg_r}="", "", IF(\'{m_name}\'!{get_column_letter(d + 12)}{16 + i}<>"", \'{m_name}\'!{get_column_letter(d + 12)}{16 + i}, ""))'
                        cell = ws.cell(row=r, column=curr_c, value=formula)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=11) # 字体适配 22.5 行高
                    else:
                        ws.cell(row=r, column=curr_c).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    ws.cell(row=r, column=curr_c).border = c_border

            heat_range = f"{get_column_letter(col_idx+1)}{row_idx+2}:{get_column_letter(col_idx+31)}{row_idx+13}"
            ws.conditional_formatting.add(heat_range, FormulaRule(formula=[f'AND({get_column_letter(col_idx+1)}{row_idx+2}<>"", {get_column_letter(col_idx+1)}{row_idx+2}<>0)'], fill=self.theme.get_fill(self.theme.SUCCESS_BG_COLOR), font=Font(color=self.theme.SUCCESS_TEXT_COLOR, bold=True, size=11), border=c_border))

    def _setup_monthly_sheets(self):
        for month_num in range(1, 13):
            ws = self.wb.create_sheet(title=f"{self.year}年{month_num}月打卡")
            num_days = calendar.monthrange(self.year, month_num)[1]; self._apply_common_settings(ws); ws.freeze_panes = f"M{self.main_table_start}"
            c_border, l_fill, d_fill, h_fill = self.theme.get_border(), self.theme.get_fill(self.theme.SUMMARY_LABEL_COLOR), self.theme.get_fill(self.theme.DASHBOARD_COLOR), self.theme.get_fill(self.theme.HEADER_COLOR)

            dash_y = self.row_offset + 1; ws.merge_cells(start_row=dash_y, start_column=3, end_row=dash_y, end_column=12)
            ws.cell(row=dash_y, column=3, value="🏆 我的坚持成就榜").font = Font(bold=True, size=16); ws.cell(row=dash_y, column=3).alignment = Alignment(horizontal='center', vertical='center')
            stat_row = dash_y + 1; ws.row_dimensions[stat_row].height = 35
            ws.merge_cells(start_row=stat_row, start_column=3, end_row=stat_row, end_column=4)
            ws.cell(row=stat_row, column=3, value="成长维度"); ws.cell(row=stat_row, column=5, value="当前状态"); ws.cell(row=stat_row, column=6, value="对比\n上月")
            for c in range(3, 7): ws.cell(row=stat_row, column=c).font = Font(bold=True, size=12); ws.cell(row=stat_row, column=c).fill = l_fill; ws.cell(row=stat_row, column=c).border = c_border; ws.cell(row=stat_row, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            helper_row = self.main_table_start + self.max_items
            for d in range(1, num_days + 1):
                col_let = get_column_letter(d + 10 + self.col_offset); ws.cell(row=helper_row, column=d + 10 + self.col_offset, value=f'=IF(COUNTA({col_let}{self.main_table_start}:{col_let}{self.main_table_start+self.max_items-1})>0, 1, 0)')
            ws.row_dimensions[helper_row].visible = False 

            labels = ["事项", "坚持事项", "平均打卡率", "平均积极率", "累计活跃天", "月度综合评分"]
            for i, label in enumerate(labels):
                r = i + 1 + stat_row; ws.row_dimensions[r].height = 45 if i == 0 else 35
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
                cell_l = ws.cell(row=r, column=3, value=label); cell_l.border = c_border; ws.cell(row=r, column=4).border = c_border; cell_l.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(i==0))
                val_c, mom_c = ws.cell(row=r, column=5), ws.cell(row=r, column=6); val_c.border = mom_c.border = c_border; val_c.alignment = mom_c.alignment = Alignment(horizontal='center', vertical='center')
                if i == 0: val_c.value = f'=COUNTIF(E{self.main_table_start}:E{self.main_table_start+self.max_items-1}, "?*")'
                elif i == 1: val_c.value = f'=COUNTIF(H{self.main_table_start}:H{self.main_table_start+self.max_items-1},">0")'
                elif i == 2: val_c.value, val_c.number_format = f'=IFERROR(AVERAGE(G{self.main_table_start}:G{self.main_table_start+self.max_items-1}),0)', '0.0%'
                elif i == 3: val_c.value, val_c.number_format = f'=IFERROR(AVERAGE(J{self.main_table_start}:J{self.main_table_start+self.max_items-1}),0)', '0.0%'
                elif i == 4: val_c.value = f'=SUM({get_column_letter(11+self.col_offset)}{helper_row}:{get_column_letter(num_days+10+self.col_offset)}{helper_row})'
                elif i == 5: 
                    # 月度综合评分逻辑：(积极率*40) + (打卡率*20) + (活跃天占比*20) + (质量比因子*20)
                    # 引用：E7=平均打卡率, E8=平均积极率, E9=累计活跃天
                    formula = f'=(E8*40) + (E7*20) + (E9/{num_days}*20) + (IFERROR(E8/E7, 0)*20)'
                    val_c.value, val_c.number_format = formula, '0.0'
                if month_num > 1:
                    prev = f"{self.year}年{month_num-1}月打卡"; curr_cell = f"E{r}"; mom_c.value = f'=IFERROR(({curr_cell}-\'{prev}\'!{curr_cell})/\'{prev}\'!{curr_cell},0)'
                else: mom_c.value = 0
                mom_c.number_format = '0.0%'

            dash_right = 11 + self.col_offset; ws.merge_cells(start_row=dash_y, start_column=dash_right, end_row=dash_y, end_column=num_days + 10 + self.col_offset)
            ws.cell(row=dash_y, column=dash_right, value="🔥 事项坚持里程碑").font = Font(bold=True, size=16); ws.cell(row=dash_y, column=dash_right).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=stat_row, column=10 + self.col_offset, value="习惯达成").fill = l_fill; ws.cell(row=stat_row, column=10 + self.col_offset).border = c_border; ws.cell(row=stat_row, column=10 + self.col_offset).alignment = Alignment(horizontal='center', vertical='center')
            
            # 顶部看板：增加对比上月的两个维度
            milestone_labels = ["事项", "已坚持", "打卡率", "积极率", "打卡对比", "积极对比"]
            for i, label in enumerate(milestone_labels):
                cell = ws.cell(row=stat_row + i + 1, column=10 + self.col_offset, value=label); cell.fill = d_fill; cell.border = c_border; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(i==0))

            for i in range(self.max_items):
                col_idx, m_r = i + 11 + self.col_offset, self.main_table_start + i
                ws.cell(row=stat_row+1, column=col_idx, value=f"=IF(E{m_r}<>\"\", E{m_r}, \"\")").alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row=stat_row+2, column=col_idx, value=f"=IF(E{m_r}<>\"\", H{m_r}, \"\")").alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=stat_row+3, column=col_idx, value=f"=IF(E{m_r}<>\"\", IFERROR(G{m_r},0), \"\")").number_format = '0.0%'; ws.cell(row=stat_row+3, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                ws.cell(row=stat_row+4, column=col_idx, value=f"=IF(E{m_r}<>\"\", IFERROR(J{m_r},0), \"\")").number_format = '0.0%'; ws.cell(row=stat_row+4, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                ws.cell(row=stat_row+5, column=col_idx, value=f"=IF(E{m_r}<>\"\", IFERROR(K{m_r},0), \"\")").number_format = '0.0%'; ws.cell(row=stat_row+5, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                ws.cell(row=stat_row+6, column=col_idx, value=f"=IF(E{m_r}<>\"\", IFERROR(L{m_r},0), \"\")").number_format = '0.0%'; ws.cell(row=stat_row+6, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

            ws.row_dimensions[self.remark_row].height = None 
            ws.merge_cells(start_row=self.remark_row, start_column=1+self.col_offset, end_row=self.remark_row, end_column=10+self.col_offset)
            for col_i in range(1+self.col_offset, 11+self.col_offset):
                cell = ws.cell(row=self.remark_row, column=col_i); cell.fill = h_fill; cell.border = c_border
                if col_i == 1+self.col_offset: cell.value = "📝 每日感悟 / 备忘录"; cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center', vertical='center')
            for d in range(1, num_days+1): ws.cell(row=self.remark_row, column=d+10+self.col_offset).border = c_border; ws.cell(row=self.remark_row, column=d+10+self.col_offset).protection = Protection(locked=False); ws.cell(row=self.remark_row, column=d+10+self.col_offset).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            ws.row_dimensions[self.main_table_start-2].height = 35; ws.row_dimensions[self.main_table_start-1].height = 35
            ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 3
            headers_cfg = [("C", "序号", 8), ("D", "类别", 12), ("E", "事项", 25), ("F", "目标\n天数", 10), ("G", "打卡率", 12), ("H", "坚持\n天数", 12), ("I", "积极\n天数", 12), ("J", "积极率", 12), ("K", "打卡\n对比", 12), ("L", "积极\n对比", 12)]
            for cl, label, width in headers_cfg:
                ws.merge_cells(f"{cl}{self.main_table_start-2}:{cl}{self.main_table_start-1}"); cell = ws[f"{cl}{self.main_table_start-2}"]; cell.value = label; cell.font = Font(bold=True, size=13); cell.fill = h_fill; cell.border = c_border; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); ws[f"{cl}{self.main_table_start-1}"].border = c_border; ws.column_dimensions[cl].width = width
            for d in range(1, num_days+1):
                col_idx = d+10+self.col_offset; dt = date(self.year, month_num, d)
                for r, val in [(self.main_table_start-2, ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][dt.weekday()]), (self.main_table_start-1, d)]:
                    c = ws.cell(row=r, column=col_idx, value=val); c.font = Font(bold=True, size=11); c.fill = h_fill; c.border = c_border; c.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(col_idx)].width = 8.5 

            dv_lock = DataValidation(type="custom", formula1=f'=$E{self.main_table_start}<>""', showErrorMessage=True, errorStyle="stop")
            dv_lock.errorTitle, dv_lock.error = "❌ 无法打卡", "该行尚未设置【事项】！请先前往事项配置页添加事项。"
            ws.add_data_validation(dv_lock); dv_lock.add(f"{get_column_letter(11+self.col_offset)}{self.main_table_start}:{get_column_letter(num_days+10+self.col_offset)}{self.main_table_start+self.max_items-1}")

            for i in range(self.max_items):
                row = self.main_table_start + i; cfg_r = i + 2; ws.row_dimensions[row].height = 40 
                for col_idx, col_let in enumerate(["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"], 1):
                    cell = ws[f"{col_let}{row}"]; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(col_let=="E"), shrink_to_fit=(col_let in ["K", "L"]))
                    if col_let == "C": cell.value = f"=IF(事项配置页!$C${cfg_r}<>\"\", 事项配置页!A{cfg_r}, \"\")"
                    elif col_let == "D": cell.value = f"=IF(事项配置页!$C${cfg_r}<>\"\", 事项配置页!B{cfg_r}, \"\")"
                    elif col_let == "E": cell.value = f"=IF(事项配置页!$C${cfg_r}<>\"\", 事项配置页!C{cfg_r}, \"\")"
                    elif col_let == "F": cell.value = f"=IF(事项配置页!$C${cfg_r}<>\"\", 事项配置页!D{cfg_r}, \"\")"
                    elif col_let == "G": cell.value = f'=IF(E{row}<>"", IFERROR(H{row}/F{row}, 0), "")'; cell.number_format = '0.0%'
                    elif col_let == "H": cell.value = f'=IF(E{row}<>"", COUNTIF({get_column_letter(11+self.col_offset)}{row}:{get_column_letter(num_days + 10 + self.col_offset)}{row},"<>" ), "")'
                    elif col_let == "I": 
                        daily_range = f"{get_column_letter(11+self.col_offset)}{row}:{get_column_letter(num_days + 10 + self.col_offset)}{row}"
                        cell.value = f'=IF(E{row}<>"", IF(事项配置页!$E${cfg_r}="", H{row}, SUMPRODUCT(--ISNUMBER(SEARCH({daily_range}, 事项配置页!$E${cfg_r}))*({daily_range}<>""))), "")'
                    elif col_let == "J": cell.value = f'=IF(E{row}<>"", IFERROR(I{row}/F{row}, 0), "")'; cell.number_format = '0.0%'
                    elif col_let == "K": # 打卡对比
                        if month_num > 1: prev = f"{self.year}年{month_num-1}月打卡"; cell.value = f'=IF(E{row}<>"", IFERROR((G{row}-\'{prev}\'!G{row})/\'{prev}\'!G{row}, 0), "")'
                        else: cell.value = f'=IF(E{row}<>"", 0, "")'
                        cell.number_format = '0.0%'
                    elif col_let == "L": # 积极对比
                        if month_num > 1: prev = f"{self.year}年{month_num-1}月打卡"; cell.value = f'=IF(E{row}<>"", IFERROR((J{row}-\'{prev}\'!J{row})/\'{prev}\'!J{row}, 0), "")'
                        else: cell.value = f'=IF(E{row}<>"", 0, "")'
                        cell.number_format = '0.0%'
                for d in range(1, num_days+1): 
                    c = ws.cell(row=row, column=d+10+self.col_offset)
                    c.protection = Protection(locked=False)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.font = Font(size=20)

            end_let = get_column_letter(num_days+10+self.col_offset); start_r, end_r = self.main_table_start, self.main_table_start + self.max_items - 1
            rate_rule = ColorScaleRule(start_type='num', start_value=0, start_color=self.theme.SCALE_RED, mid_type='num', mid_value=0.5, mid_color=self.theme.SCALE_YELLOW, end_type='num', end_value=1, end_color=self.theme.SCALE_GREEN)
            growth_rule = ColorScaleRule(start_type='num', start_value=-1, start_color=self.theme.SCALE_RED, mid_type='num', mid_value=0, mid_color=self.theme.SCALE_WHITE, end_type='num', end_value=1, end_color=self.theme.SCALE_GREEN)

            ws.conditional_formatting.add(f"C{start_r}:{end_let}{end_r}", FormulaRule(formula=[f'$E{start_r}=""'], font=Font(color="FFFFFF"), fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), border=self.theme.get_no_border(), stopIfTrue=True))
            ws.conditional_formatting.add(f"E{stat_row+3}", rate_rule); ws.conditional_formatting.add(f"E{stat_row+4}", rate_rule); ws.conditional_formatting.add(f"E{stat_row+6}", rate_rule)
            ws.conditional_formatting.add(f"F{stat_row+1}:F{stat_row+6}", growth_rule)
            dash_heat_let, dash_heat_end = get_column_letter(11+self.col_offset), get_column_letter(10+self.col_offset+self.max_items)
            ws.conditional_formatting.add(f"{dash_heat_let}{stat_row+3}:{dash_heat_end}{stat_row+3}", rate_rule)
            ws.conditional_formatting.add(f"{dash_heat_let}{stat_row+4}:{dash_heat_end}{stat_row+4}", rate_rule)
            ws.conditional_formatting.add(f"{dash_heat_let}{stat_row+5}:{dash_heat_end}{stat_row+6}", growth_rule)
            ws.conditional_formatting.add(f"G{start_r}:G{end_r}", rate_rule); ws.conditional_formatting.add(f"J{start_r}:J{end_r}", rate_rule); ws.conditional_formatting.add(f"K{start_r}:L{end_r}", growth_rule)
            ws.conditional_formatting.add(f"M{start_r}:{end_let}{end_r}", FormulaRule(formula=[f'LEN(TRIM(M{start_r}))>0'], fill=self.theme.get_fill(self.theme.SUCCESS_BG_COLOR), font=Font(color=self.theme.SUCCESS_TEXT_COLOR, bold=True, size=20), border=c_border, stopIfTrue=True))
            zebra_ranges = f"C{start_r}:F{end_r} H{start_r}:J{end_r} M{start_r}:{end_let}{end_r}"
            ws.conditional_formatting.add(zebra_ranges, FormulaRule(formula=[f'MOD(ROW()-{start_r},2)=1'], fill=self.theme.get_fill(self.theme.ZEBRA_COLOR)))
            ws.conditional_formatting.add(f"C{start_r}:{end_let}{end_r}", FormulaRule(formula=[f'$E{start_r}<>""'], border=c_border))
            ws.conditional_formatting.add(f"{dash_heat_let}{stat_row+1}:{dash_heat_end}{stat_row+6}", FormulaRule(formula=[f'{dash_heat_let}{stat_row+1}<>""'], border=c_border))
            ws.conditional_formatting.add(f"{dash_heat_let}{self.remark_row}:{end_let}{self.remark_row}", FormulaRule(formula=[f'LEN(TRIM({dash_heat_let}{self.remark_row}))>0'], fill=self.theme.get_fill(self.theme.REMARK_COLOR), stopIfTrue=True))

    def _save(self):
        self.wb.save(self.filename)
        print(f"✅ V5 默认缩放90% & 留白优化版已生成！")

if __name__ == "__main__":
    generator = HabitTrackerGenerator(year=2026)
    generator.generate()
