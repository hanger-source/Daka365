import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
import calendar
from datetime import date

# Define colors
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue background
ACHIEVEMENT_ROW_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0",
                                   fill_type="solid")  # Very light grey for achievement row
ALTERNATE_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # White for alternate rows
BORDER_STYLE = Border(left=Side(style='thin', color='C0C0C0'),  # Lighter border color
                      right=Side(style='thin', color='C0C0C0'),
                      top=Side(style='thin', color='C0C0C0'),
                      bottom=Side(style='thin', color='C0C0C0'))


def generate_365_excel_template(filename="365天打卡模板.xlsx", year=2026):
    workbook = openpyxl.Workbook()

    for month_num in range(1, 13):
        month_name = calendar.month_name[month_num]
        sheet = workbook.create_sheet(title=f"{year}年{month_num}月打卡")

        # Set default row height
        sheet.row_dimensions[1].height = 30  # Increased row height
        sheet.row_dimensions[2].height = 30  # Increased row height

        # Add "序号" column
        sheet.merge_cells('A1:A2')
        sheet['A1'] = "序号"
        sheet['A1'].font = Font(bold=True, size=13)
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A1'].fill = HEADER_FILL
        sheet['A1'].border = BORDER_STYLE
        sheet['A2'].border = BORDER_STYLE # Apply border to the merged cell's underlying cells

        # Merge B1 and B2 for "事项"
        sheet.merge_cells('B1:B2')
        sheet['B1'] = "事项"
        sheet['B1'].font = Font(bold=True, size=13)  # Increased font size
        sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['B1'].fill = HEADER_FILL
        sheet['B1'].border = BORDER_STYLE
        sheet['B2'].border = BORDER_STYLE # Apply border to the merged cell's underlying cells

        # Merge C1 and C2 for "目标"
        sheet.merge_cells('C1:C2')
        sheet['C1'] = "目标"
        sheet['C1'].font = Font(bold=True, size=13)
        sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['C1'].fill = HEADER_FILL
        sheet['C1'].border = BORDER_STYLE
        sheet['C2'].border = BORDER_STYLE # Apply border to the merged cell's underlying cells

        # Headers for Weekdays and Dates
        num_days = calendar.monthrange(year, month_num)[1]
        for col_offset in range(num_days):
            col_idx = col_offset + 4  # Start from column D (after "序号", "事项", "目标" in A, B, C)
            col_letter = get_column_letter(col_idx)

            # Weekday header (Row 1)
            day_of_week = date(year, month_num, col_offset + 1).weekday()  # 0 for Monday, 6 for Sunday
            weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            cell_weekday = sheet[f'{col_letter}1']
            cell_weekday.value = weekdays[day_of_week]
            cell_weekday.font = Font(bold=True, size=12)
            cell_weekday.alignment = Alignment(horizontal='center', vertical='center')
            cell_weekday.fill = HEADER_FILL
            cell_weekday.border = BORDER_STYLE

            # Date header (Row 2)
            cell_date = sheet[f'{col_letter}2']
            cell_date.value = col_offset + 1
            cell_date.font = Font(bold=True, size=12)
            cell_date.alignment = Alignment(horizontal='center', vertical='center')
            cell_date.fill = HEADER_FILL
            cell_date.border = BORDER_STYLE

        # Clear extra columns if month has less than 31 days
        for col_idx in range(num_days + 3,
                             31 + 3):  # From the day after the last day of the month up to (31 days max + 2 for A, B)
            col_letter = get_column_letter(col_idx)
            sheet[f'{col_letter}1'].value = ""  # Clear value
            sheet[f'{col_letter}2'].value = ""  # Clear value
            sheet[f'{col_letter}1'].border = BORDER_STYLE
            sheet[f'{col_letter}2'].border = BORDER_STYLE
            sheet[f'{col_letter}1'].fill = HEADER_FILL
            sheet[f'{col_letter}2'].fill = HEADER_FILL

        # Add "达成率" column
        achievement_rate_col_letter = get_column_letter(
            num_days + 4)  # +4 for A,B,C and 1 for the actual achievement rate column after days
        sheet[f'{achievement_rate_col_letter}1'] = "达成率"
        sheet[f'{achievement_rate_col_letter}1'].font = Font(bold=True, size=13)  # Increased font size
        sheet[f'{achievement_rate_col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'{achievement_rate_col_letter}1'].fill = HEADER_FILL
        sheet[f'{achievement_rate_col_letter}1'].border = BORDER_STYLE
        sheet[f'{achievement_rate_col_letter}2'] = ""  # Empty for the second row
        sheet[f'{achievement_rate_col_letter}2'].border = BORDER_STYLE
        sheet[f'{achievement_rate_col_letter}2'].fill = HEADER_FILL

        # Set column widths
        sheet.column_dimensions['A'].width = 8  # For "序号"
        sheet.column_dimensions['B'].width = 25  # Wider for "事项"
        sheet.column_dimensions['C'].width = 20  # Wider for "目标"
        for i in range(4, 31 + 4):  # For days and weekdays columns (D to AG)
            sheet.column_dimensions[get_column_letter(i)].width = 8  # Slightly wider
        sheet.column_dimensions[achievement_rate_col_letter].width = 18  # Wider for "达成率"

        # User guidance for emoji input
        # Note: Merging cells for guidance might conflict with individual cell borders,
        # so let's place it in a single cell and wrap text.
        guidance_row = 2 # Adjusted to remove the empty third row

        # Define the new items data with categories and items
        items_data = [
            {"序号": 1, "类别": "生活", "事项": "早睡早起"},
            {"序号": "", "类别": "", "事项": "跑步"},
            {"序号": "", "类别": "", "事项": "每天护肤"},
            {"序号": "", "类别": "", "事项": "吃早餐"},
            {"序号": 2, "类别": "学习", "事项": "睡前阅读半小时"},
            {"序号": 3, "类别": "工作", "事项": "剪视频"}
        ]

        # Populate the items based on the new data
        category_merge_info = {} # To store start_row for each category for merging
        sequence_merge_info = {} # To store start_row for each sequence for merging

        for i, item_dict in enumerate(items_data):
            row_idx = guidance_row + 1 + i  # Start from row 4
            sheet.row_dimensions[row_idx].height = 25  # Set row height for data rows

            # Fill "序号" column (A)
            if item_dict["序号"] != "":
                sheet[f'A{row_idx}'] = item_dict["序号"]
                sequence_merge_info[item_dict["序号"]] = {"start_row": row_idx, "end_row": row_idx}
            else:
                # Update end_row for the current sequence if the sequence number is empty
                last_sequence_num = max(k for k in sequence_merge_info.keys() if isinstance(k, int))
                sequence_merge_info[last_sequence_num]["end_row"] = row_idx
            
            sheet[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'A{row_idx}'].border = BORDER_STYLE

            # Fill "事项" column (B)
            if item_dict["类别"] != "":
                sheet[f'B{row_idx}'] = item_dict["类别"]
                category_merge_info[item_dict["类别"]] = {"start_row": row_idx, "end_row": row_idx}
            else:
                # Update end_row for the current category if the category is empty
                # Find the most recent non-empty category
                j = i - 1
                while j >= 0 and not items_data[j]["类别"]:
                    j -= 1
                if j >= 0:
                    last_category = items_data[j]["类别"]
                    if last_category in category_merge_info:
                        category_merge_info[last_category]["end_row"] = row_idx
            
            sheet[f'B{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'B{row_idx}'].border = BORDER_STYLE

            # Fill "目标" column (C)
            sheet[f'C{row_idx}'] = item_dict["事项"]
            sheet[f'C{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'C{row_idx}'].border = BORDER_STYLE

            # Apply border to all day cells and center alignment
            for col_offset in range(num_days):
                col_letter = get_column_letter(col_offset + 4)  # Start from D
                cell = sheet[f'{col_letter}{row_idx}']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER_STYLE

            # Apply border to achievement rate cell
            target_col = achievement_rate_col_letter
            sheet[f'{target_col}{row_idx}'].border = BORDER_STYLE

            # Achievment Rate formula - NOW COUNTING '✅' as checked
            first_day_col = get_column_letter(4)  # Column D
            last_day_col = get_column_letter(3 + num_days)
            cell = sheet[f'{target_col}{row_idx}']
            cell.value = f'=IFERROR(COUNTIF({first_day_col}{row_idx}:{last_day_col}{row_idx},"1")/(COUNTIF({first_day_col}{row_idx}:{last_day_col}{row_idx},"1")+COUNTIF({first_day_col}{row_idx}:{last_day_col}{row_idx},"0")),0)'
            cell.data_type = 'f' # Explicitly set cell type to formula
            cell.number_format = '0.00%_' # Set number format for percentage
            sheet[f'{target_col}{row_idx}'].number_format = '0.00%_'
            sheet[f'{target_col}{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells for categories in column B
        for category, info in category_merge_info.items():
            if info["start_row"] != info["end_row"]:
                sheet.merge_cells(start_row=info["start_row"], end_row=info["end_row"], start_column=2, end_column=2)
        
        # Merge cells for sequence numbers in column A
        for sequence, info in sequence_merge_info.items():
            if info["start_row"] != info["end_row"]:
                sheet.merge_cells(start_row=info["start_row"], end_row=info["end_row"], start_column=1, end_column=1)

        # Add "每日备注" row
        daily_notes_row = guidance_row + len(items_data) + 1  # Adjusted for dynamic item count
        sheet.row_dimensions[daily_notes_row].height = 30
        sheet.merge_cells(start_row=daily_notes_row, end_row=daily_notes_row, start_column=1,
                          end_column=3)  # Merge A:C for "每日备注" label
        sheet[f'A{daily_notes_row}'] = "每日备注"
        sheet[f'A{daily_notes_row}'].font = Font(bold=True, size=13)
        sheet[f'A{daily_notes_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'A{daily_notes_row}'].fill = HEADER_FILL
        sheet[f'A{daily_notes_row}'].border = BORDER_STYLE
        sheet[f'B{daily_notes_row}'].border = BORDER_STYLE
        sheet[f'B{daily_notes_row}'].fill = HEADER_FILL
        sheet[f'C{daily_notes_row}'].border = BORDER_STYLE
        sheet[f'C{daily_notes_row}'].fill = HEADER_FILL

        # Apply border to daily notes cells across the day columns
        for col_offset in range(num_days):
            col_letter = get_column_letter(col_offset + 4)  # Start from D
            sheet[f'{col_letter}{daily_notes_row}'].border = BORDER_STYLE
            sheet[f'{col_letter}{daily_notes_row}'].fill = ACHIEVEMENT_ROW_FILL

        # Apply border to the achievement rate cell for the daily notes row
        sheet[f'{achievement_rate_col_letter}{daily_notes_row}'].border = BORDER_STYLE
        sheet[f'{achievement_rate_col_letter}{daily_notes_row}'].fill = ACHIEVEMENT_ROW_FILL

        # === 月度数据汇总区域 ===
        summary_start_row = daily_notes_row + 2  # Shifted down
        sheet.merge_cells(start_row=summary_start_row, end_row=summary_start_row, start_column=1,
                          end_column=3)  # Merge A:C for summary title
        sheet[f'A{summary_start_row}'] = "月度数据汇总"
        sheet[f'A{summary_start_row}'].font = Font(bold=True, size=14)
        sheet[f'A{summary_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'A{summary_start_row}'].fill = HEADER_FILL
        sheet[f'A{summary_start_row}'].border = BORDER_STYLE
        sheet[f'B{summary_start_row}'].border = BORDER_STYLE
        sheet[f'B{summary_start_row}'].fill = HEADER_FILL
        sheet[f'C{summary_start_row}'].border = BORDER_STYLE
        sheet[f'C{summary_start_row}'].fill = HEADER_FILL
        sheet.row_dimensions[summary_start_row].height = 30

        # Define summary items and their rows
        items_count_row = summary_start_row + 1
        completed_items_row = summary_start_row + 2
        uncompleted_items_row = summary_start_row + 3
        avg_achievement_rate_row = summary_start_row + 4
        total_check_ins_row = summary_start_row + 5
        total_missed_row = summary_start_row + 6
        effective_check_in_rate_row = summary_start_row + 7

        summary_rows = [items_count_row, completed_items_row, uncompleted_items_row, avg_achievement_rate_row,
                        total_check_ins_row, total_missed_row, effective_check_in_rate_row]

        # Get the numerical index for the achievement rate column letter
        achievement_rate_col_idx = column_index_from_string(achievement_rate_col_letter)

        for r_idx in summary_rows:
            sheet.row_dimensions[r_idx].height = 25
            for c_idx in range(1, 4):  # Apply border and fill to columns A, B, C for summary items
                cell = sheet[f'{get_column_letter(c_idx)}{r_idx}']
                cell.border = BORDER_STYLE
                cell.fill = ACHIEVEMENT_ROW_FILL if (r_idx - summary_start_row) % 2 == 0 else ALTERNATE_ROW_FILL


        # 事项总数
        sheet.merge_cells(start_row=items_count_row, end_row=items_count_row, start_column=1, end_column=2) # Merge A:B for label
        sheet[f'A{items_count_row}'] = "事项总数:"
        cell = sheet[f'C{items_count_row}']
        cell.value = f'=COUNTA(B{guidance_row + 1}:B{guidance_row + len(items_data)})'
        cell.data_type = 'f'
        cell.number_format = 'General'  # Adjusted row range
        sheet[f'C{items_count_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 已完成事项数 (达成率 > 0 的事项)
        sheet.merge_cells(start_row=completed_items_row, end_row=completed_items_row, start_column=1, end_column=2) # Merge A:B for label
        sheet[f'A{completed_items_row}'] = "已完成事项数: 以达成率>0计"
        cell = sheet[f'C{completed_items_row}']
        cell.value = f'=COUNTIF({achievement_rate_col_letter}{guidance_row + 1}:{achievement_rate_col_letter}{guidance_row + len(items_data)},\">0\")'
        cell.data_type = 'f'
        cell.number_format = 'General'  # Adjusted row range
        sheet[f'C{completed_items_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 未完成事项数 (达成率 = 0 的事项)
        sheet.merge_cells(start_row=uncompleted_items_row, end_row=uncompleted_items_row, start_column=1, end_column=2) # Merge A:B for label
        sheet[f'A{uncompleted_items_row}'] = "未完成事项数: 以达成率=0计"
        cell = sheet[f'C{uncompleted_items_row}']
        cell.value = f'=COUNTIF({achievement_rate_col_letter}{guidance_row + 1}:{achievement_rate_col_letter}{guidance_row + len(items_data)},\"=0\")'
        cell.data_type = 'f'
        cell.number_format = 'General'  # Adjusted row range
        sheet[f'C{uncompleted_items_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 月度平均达成率
        sheet.merge_cells(start_row=avg_achievement_rate_row, end_row=avg_achievement_rate_row, start_column=1, end_column=2) # Merge A:B for label
        sheet[f'A{avg_achievement_rate_row}'] = "月度平均达成率:"
        cell = sheet[f'C{avg_achievement_rate_row}']
        cell.value = f'=IFERROR(AVERAGE({achievement_rate_col_letter}{guidance_row + 1}:{achievement_rate_col_letter}{guidance_row + len(items_data)}),0)'
        cell.data_type = 'f'
        cell.number_format = '0.00%_'  # Adjusted row range
        sheet[f'C{avg_achievement_rate_row}'].number_format = '0.00%_'
        sheet[f'C{avg_achievement_rate_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 总打卡天数 (所有事项打卡1的总次数)
        sheet.merge_cells(start_row=total_check_ins_row, end_row=total_check_ins_row, start_column=1, end_column=2) # Merge A:B for label
        sheet[f'A{total_check_ins_row}'] = "总打卡天数:"
        cell = sheet[f'C{total_check_ins_row}']
        cell.value = f'=SUMPRODUCT(--({get_column_letter(4)}{guidance_row + 1}:{last_day_col}{guidance_row + len(items_data)}=\"1\"))'
        cell.data_type = 'f'
        cell.number_format = 'General'  # Adjusted col and row range
        sheet[f'C{total_check_ins_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 总未打卡天数 (所有事项打卡0的总次数)
        sheet[f'A{total_missed_row}'] = "总未打卡天数:"
        cell = sheet[f'C{total_missed_row}']
        cell.value = f'=SUMPRODUCT(--({get_column_letter(4)}{guidance_row + 1}:{last_day_col}{guidance_row + len(items_data)}=\"0\"))'
        cell.data_type = 'f'
        cell.number_format = 'General'  # Adjusted col and row range
        sheet[f'C{total_missed_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 总有效打卡比例 (总打卡天数 / (总打卡天数 + 总未打卡天数))
        sheet[f'A{effective_check_in_rate_row}'] = "总有效打卡比例:"
        cell = sheet[f'C{effective_check_in_rate_row}']
        cell.value = f'=IFERROR(C{total_check_ins_row}/(C{total_check_ins_row}+C{total_missed_row}),0)'
        cell.data_type = 'f'
        cell.number_format = '0.00%_'  # Adjusted to C column
        sheet[f'C{effective_check_in_rate_row}'].number_format = '0.00%_'
        sheet[f'C{effective_check_in_rate_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for the labels of summary items

    # Remove the default "Sheet" created by openpyxl
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Add a summary sheet for YoY/MoM calculations
    summary_sheet = workbook.create_sheet(title="同比环比分析")
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 18  # Wider
    summary_sheet.column_dimensions['C'].width = 18  # Wider
    summary_sheet.column_dimensions['D'].width = 18  # Wider
    summary_sheet.column_dimensions['E'].width = 18  # Wider
    summary_sheet.column_dimensions['F'].width = 18  # For previous year's data

    summary_sheet.row_dimensions[1].height = 30  # Increased row height

    summary_sheet['A1'] = "月份"
    summary_sheet['B1'] = "月度平均达成率"
    summary_sheet['C1'] = "环比增长率"
    summary_sheet['D1'] = f"{year - 1}年平均达成率"  # Previous year for comparison
    summary_sheet['E1'] = "同比增长率"
    summary_sheet['F1'] = "注"

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        summary_sheet[f'{col_letter}1'].font = Font(bold=True, size=12)
        summary_sheet[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet[f'{col_letter}1'].fill = HEADER_FILL
        summary_sheet[f'{col_letter}1'].border = BORDER_STYLE

    # Populate summary data
    for month_num in range(1, 13):
        summary_sheet.row_dimensions[month_num + 1].height = 25  # Set row height
        summary_sheet[f'A{month_num + 1}'] = f"{year}年{month_num}月"
        summary_sheet[f'A{month_num + 1}'].border = BORDER_STYLE
        summary_sheet[f'A{month_num + 1}'].alignment = Alignment(horizontal='center', vertical='center')

        # Link to the monthly average achievement rate from the monthly sheet
        month_sheet_name = f"{year}年{month_num}月打卡"
        # We need to get the column letter for the average achievement rate dynamically
        num_days_in_month = calendar.monthrange(year, month_num)[1]
        ach_rate_col_monthly_sheet = get_column_letter(num_days_in_month + 4)  # Adjust for new A,B,C columns
        summary_sheet[
            f'B{month_num + 1}'] = f'=\'{month_sheet_name}\'!${ach_rate_col_monthly_sheet}${avg_achievement_rate_row}'  # Link to the new average row
        summary_sheet[f'B{month_num + 1}'].number_format = '0.00%_'
        summary_sheet[f'B{month_num + 1}'].border = BORDER_STYLE
        summary_sheet[f'B{month_num + 1}'].alignment = Alignment(horizontal='center', vertical='center')

        # MoM (Month-over-Month) growth rate
        if month_num > 1:
            summary_sheet[f'C{month_num + 1}'] = f'=IFERROR((B{month_num + 1}-B{month_num})/B{month_num},0)'
        else:
            summary_sheet[f'C{month_num + 1}'] = ""
        summary_sheet[f'C{month_num + 1}'].number_format = '0.00%_'
        summary_sheet[f'C{month_num + 1}'].border = BORDER_STYLE
        summary_sheet[f'C{month_num + 1}'].alignment = Alignment(horizontal='center', vertical='center')

        # Previous year's data for YoY (Manual input)
        summary_sheet[f'D{month_num + 1}'] = ""  # Placeholder for manual input
        summary_sheet[f'D{month_num + 1}'].number_format = '0.00%_'
        summary_sheet[f'D{month_num + 1}'].border = BORDER_STYLE
        summary_sheet[f'D{month_num + 1}'].alignment = Alignment(horizontal='center', vertical='center')

        # YoY (Year-over-Year) growth rate
        summary_sheet[f'E{month_num + 1}'] = f'=IFERROR((B{month_num + 1}-D{month_num + 1})/D{month_num + 1},0)'
        summary_sheet[f'E{month_num + 1}'].number_format = '0.00%_'
        summary_sheet[f'E{month_num + 1}'].border = BORDER_STYLE
        summary_sheet[f'E{month_num + 1}'].alignment = Alignment(horizontal='center', vertical='center')

        summary_sheet[f'F{month_num + 1}'] = "D列手动填写上一年数据（如月度平均达成率）进行同比计算"
        summary_sheet[f'F{month_num + 1}'].font = Font(size=9, color='808080')
        summary_sheet[f'F{month_num + 1}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        summary_sheet[f'F{month_num + 1}'].border = BORDER_STYLE

    workbook.save(filename)
    print(f"Excel模板 \'{filename}\' 已生成。")


if __name__ == "__main__":
    generate_365_excel_template()