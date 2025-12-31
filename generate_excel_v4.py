import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
import calendar
from datetime import date

# ==========================================
# 🎨 全局样式与色系配置 (方便后期一键换肤)
# ==========================================
TEXT_COLOR = "000000"  # 通用文字颜色

# --- 边框与背景色系 ---
BORDER_COLOR = "94A3B8"  # 边框颜色 (清透蓝灰)
CLEAR_BORDER_SIDE = Side(style='thin', color=BORDER_COLOR)
BORDER_STYLE = Border(left=CLEAR_BORDER_SIDE, right=CLEAR_BORDER_SIDE, top=CLEAR_BORDER_SIDE, bottom=CLEAR_BORDER_SIDE)
NO_BORDER = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

DASHBOARD_COLOR = "D8E2DC"      # 看板次级背景 (灰岩绿)
SUMMARY_LABEL_COLOR = "FFE5D9"  # 核心维度背景 (蜜粉色)
HEADER_COLOR = "ECE4DB"         # 表头背景 (奶油杏)
ZEBRA_COLOR = "F8FAFC"          # 斑马纹背景 (极淡蓝灰)
REMARK_COLOR = "F1F5F9"         # 备注输入背景 (云朵蓝)

# --- 状态指示色系 ---
SUCCESS_BG_COLOR = "D1FAE5"     # 打卡成功背景 (薄荷曼波)
SUCCESS_TEXT_COLOR = "059669"   # 打卡成功文字 (深草绿)
ERROR_BG_COLOR = "FEE2E2"       # 错误/重复提示背景 (浅淡红)
ERROR_TEXT_COLOR = "B91C1C"     # 错误/重复提示文字 (深砖红)

# --- 色阶定义 (红-黄-绿 渐变) ---
SCALE_RED = "FCA5A5"            # 警告红
SCALE_YELLOW = "FDE68A"         # 过渡黄
SCALE_GREEN = "86EFAC"          # 达成绿
SCALE_WHITE = "FFFFFF"          # 纯白 (用于环比中点)

# --- 样式对象初始化 ---
DASHBOARD_FILL = PatternFill(start_color=DASHBOARD_COLOR, end_color=DASHBOARD_COLOR, fill_type="solid")
SUMMARY_LABEL_FILL = PatternFill(start_color=SUMMARY_LABEL_COLOR, end_color=SUMMARY_LABEL_COLOR, fill_type="solid")
HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
ZEBRA_FILL = PatternFill(start_color=ZEBRA_COLOR, end_color=ZEBRA_COLOR, fill_type="solid")
SUCCESS_FILL = PatternFill(start_color=SUCCESS_BG_COLOR, end_color=SUCCESS_BG_COLOR, fill_type="solid") 
REMARK_FILL = PatternFill(start_color=REMARK_COLOR, end_color=REMARK_COLOR, fill_type="solid")
ERROR_FILL = PatternFill(start_color=ERROR_BG_COLOR, end_color=ERROR_BG_COLOR, fill_type="solid")

SUCCESS_FONT = Font(color=SUCCESS_TEXT_COLOR, bold=True)
ERROR_FONT = Font(color=ERROR_TEXT_COLOR, bold=True)
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF")

# ==========================================

def generate_perfect_365_excel(filename="365天打卡模板_v4_顶置看板版.xlsx", year=2026, max_items=50):
    workbook = openpyxl.Workbook()
    
    # --- 1. 配置页 ---
    config_sheet = workbook.active
    config_sheet.title = "配置页"
    config_sheet.protection.set_password('123456'); config_sheet.protection.sheet = True
    config_sheet.protection.formatCells = False; config_sheet.protection.insertRows = False; config_sheet.protection.deleteRows = False; config_sheet.protection.sort = False; config_sheet.protection.autoFilter = False
    config_sheet.sheet_view.showGridLines = False

    config_headers = ["序号", "类别", "习惯事项", "目标天数"]
    config_sheet.row_dimensions[1].height = 40
    for i, h in enumerate(config_headers, 1):
        cell = config_sheet.cell(row=1, column=i, value=h); cell.font = Font(bold=True, size=14, color=TEXT_COLOR); cell.fill = DASHBOARD_FILL; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = BORDER_STYLE; cell.protection = Protection(locked=True) 
    
    for r in range(2, max_items + 2):
        config_sheet.row_dimensions[r].height = 35 
        for c in range(1, 27): 
            cell = config_sheet.cell(row=r, column=c); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(c==3)); cell.font = Font(size=12, color=TEXT_COLOR); cell.protection = Protection(locked=False)
            if c == 1: cell.value = f'=IF(C{r}<>"", ROW()-1, "")'

    config_item_range = f"C2:C{max_items + 1}"
    config_dv = DataValidation(type="custom", formula1=f'COUNTIF($C$2:$C${max_items+1}, C2)<=1', showErrorMessage=True, errorStyle="stop")
    config_dv.errorTitle = "❌ 习惯事项重复"; config_dv.error = "该习惯事项已经存在！请勿重复添加。"; config_sheet.add_data_validation(config_dv); config_dv.add(config_item_range)

    config_range = f"A2:D{max_items + 1}"
    config_sheet.conditional_formatting.add(config_range, FormulaRule(formula=[f'COUNTIF($C$2:$C${max_items+1}, C2)>1'], fill=ERROR_FILL, font=ERROR_FONT))
    config_sheet.conditional_formatting.add(config_range, FormulaRule(formula=['LEN(TRIM($C2))>0'], border=BORDER_STYLE))
    config_sheet.conditional_formatting.add(config_range, FormulaRule(formula=['AND(LEN(TRIM($C2))>0, MOD(ROW(),2)=1)'], fill=ZEBRA_FILL))

    initial_items = [["生活", "早睡早起", 21], ["生活", "跑步", 21], ["学习", "睡前阅读", 10]]
    for r_idx, row_data in enumerate(initial_items, 2):
        for c_idx, val in enumerate(row_data, 2): config_sheet.cell(row=r_idx, column=c_idx, value=val)
    config_sheet.column_dimensions['C'].width = 45
    config_sheet.column_dimensions['D'].width = 15

    # --- 2. 12个月份页 ---
    ROW_OFFSET, COL_OFFSET = 2, 2
    REMARK_ROW, HEADER_START, MAIN_TABLE_START = 11 + ROW_OFFSET, 12 + ROW_OFFSET, 14 + ROW_OFFSET 
    
    for month_num in range(1, 13):
        sheet = workbook.create_sheet(title=f"{year}年{month_num}月打卡")
        num_days = calendar.monthrange(year, month_num)[1]
        sheet.protection.sheet = True; sheet.protection.password = '123456'
        sheet.freeze_panes = f"J{MAIN_TABLE_START}"; sheet.sheet_view.showGridLines = False

        # --- A. 仪表盘 ---
        dash_left_col = 3
        sheet.merge_cells(start_row=1 + ROW_OFFSET, start_column=dash_left_col, end_row=1 + ROW_OFFSET, end_column=dash_left_col + 6)
        title_dash = sheet.cell(row=1 + ROW_OFFSET, column=dash_left_col); title_dash.value = "🏆 我的自律成就榜"; title_dash.font = Font(bold=True, size=16, color=TEXT_COLOR); title_dash.alignment = Alignment(horizontal='left', vertical='center')
        
        stat_row = 2 + ROW_OFFSET
        sheet.row_dimensions[stat_row].height = 35 
        sheet.merge_cells(start_row=stat_row, start_column=dash_left_col, end_row=stat_row, end_column=dash_left_col + 1)
        sheet.cell(row=stat_row, column=dash_left_col, value="成长维度")
        sheet.cell(row=stat_row, column=dash_left_col + 2, value="当前状态"); sheet.cell(row=stat_row, column=dash_left_col + 3, value="对比\n上月")
        for c in range(dash_left_col, dash_left_col + 4):
            cell = sheet.cell(row=stat_row, column=c); cell.font = Font(bold=True, size=12, color=TEXT_COLOR); cell.fill = SUMMARY_LABEL_FILL; cell.border = BORDER_STYLE; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        labels = ["习惯事项", "坚持事项", "平均达成率", "累计打卡天", "总体完成率"]
        for i, label in enumerate(labels):
            r = i + 3 + ROW_OFFSET
            sheet.row_dimensions[r].height = 45 if i == 0 else 35 
            sheet.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            cell_label = sheet.cell(row=r, column=3, value=label); cell_label.border = BORDER_STYLE; sheet.cell(row=r, column=4).border = BORDER_STYLE
            cell_label.font = Font(size=11, color=TEXT_COLOR); cell_label.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(i==0))
            
            val_c, mom_c = sheet.cell(row=r, column=dash_left_col + 2), sheet.cell(row=r, column=dash_left_col + 3)
            val_c.border = mom_c.border = BORDER_STYLE; val_c.font = mom_c.font = Font(color=TEXT_COLOR); val_c.alignment = mom_c.alignment = Alignment(horizontal='center', vertical='center')
            if i == 0: val_c.value = f'=COUNTIF(E{MAIN_TABLE_START}:E{MAIN_TABLE_START+max_items-1}, "?*")'
            elif i == 1: val_c.value = f'=COUNTIF(G{MAIN_TABLE_START}:G{MAIN_TABLE_START+max_items-1},">0")'
            elif i == 2: val_c.value, val_c.number_format = f'=IFERROR(AVERAGE(G{MAIN_TABLE_START}:G{MAIN_TABLE_START+max_items-1}),0)', '0.0%'
            elif i == 3: val_c.value = f'=SUM(H{MAIN_TABLE_START}:H{MAIN_TABLE_START+max_items-1})'
            elif i == 4: val_c.value, val_c.number_format = f'=IFERROR({get_column_letter(dash_left_col+2)}{r-1}/({get_column_letter(dash_left_col+2)}{3+ROW_OFFSET}*{num_days}),0)', '0.0%'
            if month_num > 1:
                prev = f"{year}年{month_num-1}月打卡"; curr_cell = f"{get_column_letter(dash_left_col+2)}{r}"; mom_c.value = f'=IFERROR(({curr_cell}-\'{prev}\'!{curr_cell})/\'{prev}\'!{curr_cell},0)'
            else: mom_c.value = 0
            mom_c.number_format = '0.0%'

        # 习惯看板 (习惯达成看板)
        dash_right_col_start = 8 + COL_OFFSET; dash_right_col_end = num_days + 7 + COL_OFFSET
        sheet.merge_cells(start_row=1 + ROW_OFFSET, start_column=dash_right_col_start, end_row=1 + ROW_OFFSET, end_column=dash_right_col_end)
        title_items = sheet.cell(row=1 + ROW_OFFSET, column=dash_right_col_start, value="🔥 习惯进化里程碑"); title_items.font = Font(bold=True, size=16, color=TEXT_COLOR); title_items.alignment = Alignment(horizontal='center', vertical='center')
        
        label_col = 7 + COL_OFFSET
        cell_core = sheet.cell(row=2 + ROW_OFFSET, column=label_col, value="习惯达成")
        cell_core.font = Font(bold=True, size=11); cell_core.fill = SUMMARY_LABEL_FILL; cell_core.border = BORDER_STYLE; cell_core.alignment = Alignment(horizontal='center', vertical='center')
        
        for i, label in enumerate(["习惯事项", "已坚持", "达成率", "对比上月"]):
            r = i + 3 + ROW_OFFSET
            cell_h = sheet.cell(row=r, column=label_col, value=label)
            cell_h.font = Font(size=10, color=TEXT_COLOR); cell_h.fill = DASHBOARD_FILL; cell_h.border = BORDER_STYLE; cell_h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(i==0))
        
        for i in range(max_items):
            col_idx = i + 8 + COL_OFFSET; main_row = MAIN_TABLE_START + i; sheet.cell(row=3 + ROW_OFFSET, column=col_idx, value=f"=IF(E{main_row}<>\"\", E{main_row}, \"\")")
            sheet.cell(row=4 + ROW_OFFSET, column=col_idx, value=f"=IF(E{main_row}<>\"\", H{main_row}, \"\")"); sheet.cell(row=5 + ROW_OFFSET, column=col_idx, value=f"=IF(E{main_row}<>\"\", G{main_row}, \"\")"); sheet.cell(row=6 + ROW_OFFSET, column=col_idx, value=f"=IF(E{main_row}<>\"\", I{main_row}, \"\")")
            for r in range(3 + ROW_OFFSET, 7 + ROW_OFFSET):
                cell = sheet.cell(row=r, column=col_idx); cell.font = Font(size=10, color=TEXT_COLOR); 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(r==3+ROW_OFFSET))
                if r in [5 + ROW_OFFSET, 6 + ROW_OFFSET]: cell.number_format = '0.0%'

        # --- B. 每日感悟 ---
        sheet.row_dimensions[REMARK_ROW].height = None 
        sheet.merge_cells(start_row=REMARK_ROW, start_column=1 + COL_OFFSET, end_row=REMARK_ROW, end_column=7 + COL_OFFSET)
        remark_title = sheet.cell(row=REMARK_ROW, column=1 + COL_OFFSET); remark_title.value = "📝 每日感悟 / 备忘录"; remark_title.font = Font(bold=True, size=12, color=TEXT_COLOR); remark_title.fill = HEADER_FILL; remark_title.alignment = Alignment(horizontal='center', vertical='center'); remark_title.border = BORDER_STYLE
        for c in range(1 + COL_OFFSET, 8 + COL_OFFSET): sheet.cell(row=REMARK_ROW, column=c).border = BORDER_STYLE
        for d in range(1, num_days + 1):
            cell = sheet.cell(row=REMARK_ROW, column=d + 7 + COL_OFFSET); cell.border = BORDER_STYLE; cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True); cell.font = Font(size=10, color=TEXT_COLOR); cell.protection = Protection(locked=False)

        # --- C. 表头 ---
        sheet.row_dimensions[HEADER_START].height = 35; sheet.row_dimensions[HEADER_START + 1].height = 35
        main_headers = [("C", "序号"), ("D", "类别"), ("E", "习惯事项"), ("F", "目标\n天数"), ("G", "达成率"), ("H", "坚持\n天数"), ("I", "对比\n上月")]
        for col_let, label in main_headers:
            sheet.merge_cells(f"{col_let}{HEADER_START}:{col_let}{HEADER_START+1}"); cell = sheet[f"{col_let}{HEADER_START}"]; cell.value = label; cell.font = Font(bold=True, size=13, color=TEXT_COLOR); cell.fill = HEADER_FILL; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = BORDER_STYLE; sheet[f"{col_let}{HEADER_START+1}"].border = BORDER_STYLE
        weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        for d in range(1, num_days + 1):
            col_idx = d + 7 + COL_OFFSET; day_of_week = date(year, month_num, d).weekday()
            for r, val in [(HEADER_START, weekdays[day_of_week]), (HEADER_START + 1, d)]:
                cell = sheet.cell(row=r, column=col_idx, value=val); cell.font = Font(bold=True, size=11, color=TEXT_COLOR); cell.fill = HEADER_FILL; cell.border = BORDER_STYLE; cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- D. 数据行 ---
        checkin_area_range = f"{get_column_letter(8+COL_OFFSET)}{MAIN_TABLE_START}:{get_column_letter(num_days + 7 + COL_OFFSET)}{MAIN_TABLE_START + max_items - 1}"
        
        dv_lock = DataValidation(type="custom", formula1=f'=$E{MAIN_TABLE_START}<>""', showErrorMessage=True, errorStyle="stop")
        dv_lock.errorTitle = "❌ 无法打卡"; dv_lock.error = "请先设置【习惯事项】后再进行打卡！"; sheet.add_data_validation(dv_lock); dv_lock.add(checkin_area_range)

        for i in range(max_items):
            row = MAIN_TABLE_START + i; cfg_r = i + 2; 
            sheet.row_dimensions[row].height = 40 
            for col_idx, col_let in enumerate(["C", "D", "E", "F", "G", "H", "I"], 1):
                cell = sheet[f"{col_let}{row}"]
                if col_let == "C": cell.value = f"=IF(配置页!$C${cfg_r}<>\"\", 配置页!A{cfg_r}, \"\")"
                elif col_let == "D": cell.value = f"=IF(配置页!$C${cfg_r}<>\"\", 配置页!B{cfg_r}, \"\")"
                elif col_let == "E": cell.value = f"=IF(配置页!$C${cfg_r}<>\"\", 配置页!C{cfg_r}, \"\")"
                elif col_let == "F": cell.value = f"=IF(配置页!$C${cfg_r}<>\"\", 配置页!D{cfg_r}, \"\")"
                elif col_let == "G": cell.value = f'=IF(E{row}<>"", IFERROR(H{row}/F{row}, 0), "")'; cell.number_format = '0.0%'
                elif col_let == "H": cell.value = f'=IF(E{row}<>"", COUNTIF({get_column_letter(8+COL_OFFSET)}{row}:{get_column_letter(num_days + 7 + COL_OFFSET)}{row},"<>" ), "")'
                elif col_let == "I":
                    if month_num > 1: prev = f"{year}年{month_num-1}月打卡"; cell.value = f'=IF(E{row}<>"", IFERROR((G{row}-\'{prev}\'!G{row})/\'{prev}\'!G{row}, 0), "")'
                    else: cell.value = f'=IF(E{row}<>"", 0, "")'
                    cell.number_format = '0.0%'
                cell.font = Font(color=TEXT_COLOR); 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(col_let == "E"))
                cell.border = BORDER_STYLE
            for d in range(1, num_days + 1):
                cell = sheet.cell(row=row, column=d + 7 + COL_OFFSET); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.font = Font(size=12, color=TEXT_COLOR); cell.protection = Protection(locked=False)

        # --- 💡 E. 条件格式 (优先级重构) ---
        table_full_range = f"C{MAIN_TABLE_START}:{get_column_letter(num_days + 7 + COL_OFFSET)}{MAIN_TABLE_START + max_items - 1}"
        dash_total_end = max_items + 7 + COL_OFFSET
        dash_board_range = f"{get_column_letter(8+COL_OFFSET)}{3+ROW_OFFSET}:{get_column_letter(dash_total_end)}{6+ROW_OFFSET}"

        # 1. 空行绝对过滤 (隐形粘贴内容)
        sheet.conditional_formatting.add(table_full_range, FormulaRule(formula=[f'$E{MAIN_TABLE_START}=""'], font=WHITE_FONT, fill=WHITE_FILL, border=NO_BORDER, stopIfTrue=True))
        # 2. 打卡变色 (带边框)
        sheet.conditional_formatting.add(checkin_area_range, FormulaRule(formula=[f'LEN(TRIM({get_column_letter(8+COL_OFFSET)}{MAIN_TABLE_START}))>0'], fill=SUCCESS_FILL, font=SUCCESS_FONT, border=BORDER_STYLE, stopIfTrue=True))
        # 3. 斑马纹
        sheet.conditional_formatting.add(table_full_range, FormulaRule(formula=[f'AND($E{MAIN_TABLE_START}<>"", MOD(ROW()-{MAIN_TABLE_START},2)=1)'], fill=ZEBRA_FILL))
        # 4. 动态边框
        sheet.conditional_formatting.add(table_full_range, FormulaRule(formula=[f'$E{MAIN_TABLE_START}<>""'], border=BORDER_STYLE))
        sheet.conditional_formatting.add(dash_board_range, FormulaRule(formula=[f'LEN(TRIM({get_column_letter(8+COL_OFFSET)}{3+ROW_OFFSET}))>0'], border=BORDER_STYLE))

        # 色阶规则 (使用顶部定义的颜色)
        rate_rule = ColorScaleRule(start_type='num', start_value=0, start_color=SCALE_RED, mid_type='num', mid_value=0.5, mid_color=SCALE_YELLOW, end_type='num', end_value=1, end_color=SCALE_GREEN)
        growth_rule = ColorScaleRule(start_type='num', start_value=-1, start_color=SCALE_RED, mid_type='num', mid_value=0, mid_color=SCALE_WHITE, end_type='num', end_value=1, end_color=SCALE_GREEN)
        
        sheet.conditional_formatting.add(f"G{MAIN_TABLE_START}:G{MAIN_TABLE_START+max_items-1}", rate_rule); sheet.conditional_formatting.add(f"I{MAIN_TABLE_START}:I{MAIN_TABLE_START+max_items-1}", growth_rule)
        sheet.conditional_formatting.add(f"{get_column_letter(8+COL_OFFSET)}{5+ROW_OFFSET}:{get_column_letter(num_days+7+COL_OFFSET)}{5+ROW_OFFSET}", rate_rule); sheet.conditional_formatting.add(f"{get_column_letter(8+COL_OFFSET)}{6+ROW_OFFSET}:{get_column_letter(num_days+7+COL_OFFSET)}{6+ROW_OFFSET}", growth_rule)
        sheet.conditional_formatting.add(f"{get_column_letter(dash_left_col+2)}{5+ROW_OFFSET}", rate_rule); sheet.conditional_formatting.add(f"{get_column_letter(dash_left_col+2)}{7+ROW_OFFSET}", rate_rule); sheet.conditional_formatting.add(f"{get_column_letter(dash_left_col+3)}{3+ROW_OFFSET}:{get_column_letter(dash_left_col+3)}{7+ROW_OFFSET}", growth_rule)
        sheet.conditional_formatting.add(f"{get_column_letter(8+COL_OFFSET)}{REMARK_ROW}:{get_column_letter(num_days + 7 + COL_OFFSET)}{REMARK_ROW}", FormulaRule(formula=[f'LEN(TRIM({get_column_letter(8+COL_OFFSET)}{REMARK_ROW}))>0'], fill=REMARK_FILL, stopIfTrue=True))

        sheet.column_dimensions['A'].width = 5; sheet.column_dimensions['B'].width = 5
        for col, width in [('C', 8), ('D', 12), ('E', 25), ('F', 10), ('G', 9), ('H', 9), ('I', 12)]: sheet.column_dimensions[col].width = width
        for d in range(1, num_days + 1): sheet.column_dimensions[get_column_letter(d + 7 + COL_OFFSET)].width = 7

    workbook.save(filename)
    print(f"最终视觉修复版 V4.81 '{filename}' 已生成！")

if __name__ == "__main__":
    generate_perfect_365_excel()
