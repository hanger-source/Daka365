import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import calendar
from datetime import date

# 定义样式
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ACHIEVEMENT_ROW_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
ALTERNATE_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BORDER_STYLE = Border(left=Side(style='thin', color='C0C0C0'),
                      right=Side(style='thin', color='C0C0C0'),
                      top=Side(style='thin', color='C0C0C0'),
                      bottom=Side(style='thin', color='C0C0C0'))

def generate_365_excel_template(filename="365天打卡模板_v2_2026.xlsx", year=2026):
    workbook = openpyxl.Workbook()

    # 预定义事项数据
    items_data = [
        {"序号": 1, "类别": "生活", "事项": "早睡早起", "目标": 21},
        {"序号": 2, "类别": "", "事项": "跑步", "目标": 21},
        {"序号": 3, "类别": "", "事项": "每天护肤", "目标": 31},
        {"序号": 4, "类别": "", "事项": "吃早餐", "目标": 31},
        {"序号": 5, "类别": "学习", "事项": "睡前阅读半小时", "目标": 10},
        {"序号": 6, "类别": "工作", "事项": "剪视频", "目标": 14}
    ]

    for month_num in range(1, 13):
        sheet_title = f"{year}年{month_num}月打卡"
        sheet = workbook.create_sheet(title=sheet_title)

        num_days = calendar.monthrange(year, month_num)[1]
        ach_col_idx = num_days + 5  
        mom_col_idx = num_days + 6  
        ach_col_letter = get_column_letter(ach_col_idx)
        mom_col_letter = get_column_letter(mom_col_idx)

        # 设置表头行高
        sheet.row_dimensions[1].height = 35
        sheet.row_dimensions[2].height = 35

        # 1. 顶部表头 (A, B, C, D)
        headers = [("A1", "序号"), ("B1", "类别"), ("C1", "事项"), ("D1", "目标")]
        for cell_ref, label in headers:
            sheet.merge_cells(f"{cell_ref}:{cell_ref[0]}2")
            cell = sheet[cell_ref]
            cell.value = label
            cell.font = Font(bold=True, size=14) # 增加字体大小
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = HEADER_FILL
            cell.border = BORDER_STYLE
            sheet[f"{cell_ref[0]}2"].border = BORDER_STYLE

        # 2. 日期表头 (E列开始)
        for d in range(1, num_days + 1):
            col_letter = get_column_letter(d + 4)
            day_of_week = date(year, month_num, d).weekday()
            weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            
            # 星期
            sheet[f"{col_letter}1"].value = weekdays[day_of_week]
            sheet[f"{col_letter}1"].font = Font(bold=True, size=12)
            sheet[f"{col_letter}1"].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f"{col_letter}1"].fill = HEADER_FILL
            sheet[f"{col_letter}1"].border = BORDER_STYLE
            
            # 日期
            sheet[f"{col_letter}2"].value = d
            sheet[f"{col_letter}2"].font = Font(bold=True, size=12)
            sheet[f"{col_letter}2"].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f"{col_letter}2"].fill = HEADER_FILL
            sheet[f"{col_letter}2"].border = BORDER_STYLE

        # 3. 指标表头 (右侧)
        for idx, label in [(ach_col_idx, "达成率"), (mom_col_idx, "环比增长率")]:
            col_letter = get_column_letter(idx)
            sheet.merge_cells(f"{col_letter}1:{col_letter}2")
            cell = sheet[f"{col_letter}1"]
            cell.value = label
            cell.font = Font(bold=True, size=14) # 增加字体大小
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = HEADER_FILL
            cell.border = BORDER_STYLE
            sheet[f"{col_letter}2"].border = BORDER_STYLE
            sheet.column_dimensions[col_letter].width = 18

        # 4. 事项内容填充
        category_merge_info = {}
        start_row, end_row = 3, len(items_data) + 2
        col_e_start, col_e_end = get_column_letter(5), get_column_letter(num_days + 4)
        
        for i, item in enumerate(items_data):
            row = i + 3
            sheet.row_dimensions[row].height = 30 # 增加数据行高
            sheet[f"A{row}"], sheet[f"B{row}"], sheet[f"C{row}"], sheet[f"D{row}"] = item["序号"], item["类别"], item["事项"], item["目标"]
            
            for col in ["A", "B", "C", "D"]:
                sheet[f"{col}{row}"].font = Font(size=12) # 增加内容字体
                sheet[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f"{col}{row}"].border = BORDER_STYLE

            if item["类别"]: category_merge_info[item["类别"]] = {"start": row, "end": row}
            elif category_merge_info: category_merge_info[next(reversed(category_merge_info))]["end"] = row

            for d_col in range(5, num_days + 5):
                cell = sheet.cell(row=row, column=d_col)
                cell.border = BORDER_STYLE
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 达成率
            ach_c = sheet[f"{ach_col_letter}{row}"]
            ach_c.value = f'=IFERROR(COUNTIF({col_e_start}{row}:{col_e_end}{row},"<>" )/D{row},0)'
            ach_c.data_type, ach_c.number_format = 'f', '0.00%'
            ach_c.font = Font(size=12)
            ach_c.alignment, ach_c.border = Alignment(horizontal='center', vertical='center'), BORDER_STYLE

            # 环比
            mom_c = sheet[f"{mom_col_letter}{row}"]
            if month_num > 1:
                prev = f"{year}年{month_num-1}月打卡"
                prev_ach_col = get_column_letter(calendar.monthrange(year, month_num-1)[1] + 5)
                mom_c.value = f'=IFERROR(({ach_col_letter}{row}-\'{prev}\'!{prev_ach_col}{row})/\'{prev}\'!{prev_ach_col}{row},0)'
                mom_c.data_type = 'f'
            mom_c.number_format, mom_c.alignment, mom_c.border = '0.00%', Alignment(horizontal='center', vertical='center'), BORDER_STYLE
            mom_c.font = Font(size=12)

        for info in category_merge_info.values():
            if info["start"] != info["end"]: sheet.merge_cells(start_row=info["start"], end_row=info["end"], start_column=2, end_column=2)

        # 5. 每日备注
        note_row = end_row + 1
        sheet.row_dimensions[note_row].height = 35 # 增加高度
        sheet.merge_cells(f"A{note_row}:D{note_row}")
        sheet[f"A{note_row}"] = "每日备注"
        sheet[f"A{note_row}"].font, sheet[f"A{note_row}"].fill = Font(bold=True, size=14), HEADER_FILL
        sheet[f"A{note_row}"].alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, mom_col_idx + 1):
            cell = sheet.cell(row=note_row, column=c)
            cell.border = BORDER_STYLE
            if c > 4: cell.fill = ACHIEVEMENT_ROW_FILL

        # 6. 月度数据汇总区域 (全自动、细分)
        summary_start = note_row + 2
        
        # --- 1. 汇总大标题 ---
        sheet.row_dimensions[summary_start].height = 40 # 增加大标题行高
        sheet.merge_cells(start_row=summary_start, end_row=summary_start, start_column=1, end_column=4)
        title_c = sheet[f"A{summary_start}"]
        title_c.value, title_c.font, title_c.fill = "月度数据汇总", Font(bold=True, size=16), HEADER_FILL
        title_c.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, 5): sheet.cell(row=summary_start, column=c).border = BORDER_STYLE

        # --- 2. 事项明细部分 ---
        item_header_row = summary_start + 1
        sheet.row_dimensions[item_header_row].height = 30
        sheet.merge_cells(f"A{item_header_row}:B{item_header_row}")
        sheet[f"A{item_header_row}"] = "事项明细"
        sheet[f"C{item_header_row}"] = "打卡天数"
        sheet[f"D{item_header_row}"] = "达成率"
        for c in range(1, 5):
            cell = sheet.cell(row=item_header_row, column=c)
            cell.font, cell.fill, cell.border = Font(bold=True, size=13), ACHIEVEMENT_ROW_FILL, BORDER_STYLE
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充每个事项的细分汇总
        current_r = item_header_row + 1
        for i, item in enumerate(items_data):
            row_in_data = i + 3
            sheet.row_dimensions[current_r].height = 30
            sheet.merge_cells(f"A{current_r}:B{current_r}")
            sheet[f"A{current_r}"] = item["事项"]
            sheet[f"A{current_r}"].font = Font(size=12)
            
            # 本月打卡天数
            days_c = sheet[f"C{current_r}"]
            days_c.value = f'=COUNTIF({col_e_start}{row_in_data}:{col_e_end}{row_in_data},"<>" )'
            days_c.font = Font(size=12)
            
            # 达成率
            sum_ach_c = sheet[f"D{current_r}"]
            sum_ach_c.value = f'={ach_col_letter}{row_in_data}'
            sum_ach_c.number_format = '0.00%'
            sum_ach_c.font = Font(size=12)
            
            for c in range(1, 5):
                cell = sheet.cell(row=current_r, column=c)
                cell.border = BORDER_STYLE
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = ALTERNATE_ROW_FILL if i % 2 != 0 else ACHIEVEMENT_ROW_FILL
            current_r += 1

        # --- 3. 总体统计部分 ---
        total_header_row = current_r
        sheet.row_dimensions[total_header_row].height = 30
        sheet.merge_cells(f"A{total_header_row}:B{total_header_row}")
        sheet[f"A{total_header_row}"] = "统计指标"
        sheet[f"C{total_header_row}"] = "数值"
        sheet[f"D{total_header_row}"] = "环比增长"
        for c in range(1, 5):
            cell = sheet.cell(row=total_header_row, column=c)
            cell.font, cell.fill, cell.border = Font(bold=True, size=13), ACHIEVEMENT_ROW_FILL, BORDER_STYLE
            cell.alignment = Alignment(horizontal='center', vertical='center')

        labels = ["事项总数:", "已完成事项(达成率>0):", "月度平均达成率:", "总打卡天数:", "总有效打卡比例:"]
        for i, label in enumerate(labels):
            r = total_header_row + 1 + i
            sheet.row_dimensions[r].height = 30
            sheet.merge_cells(f"A{r}:B{r}")
            sheet[f"A{r}"] = label
            sheet[f"A{r}"].font = Font(size=12)
            fill = ACHIEVEMENT_ROW_FILL if i % 2 == 0 else ALTERNATE_ROW_FILL
            
            for c in range(1, 5):
                cell = sheet.cell(row=r, column=c)
                cell.border, cell.fill = BORDER_STYLE, fill
            
            val_c, sum_mom_c = sheet[f"C{r}"], sheet[f"D{r}"]
            val_c.font = Font(size=12)
            sum_mom_c.font = Font(size=12)
            val_c.data_type, val_c.alignment = 'f', Alignment(horizontal='center', vertical='center')
            sum_mom_c.data_type, sum_mom_c.number_format, sum_mom_c.alignment = 'f', '0.00%', Alignment(horizontal='center', vertical='center')

            if i == 0: val_c.value = f'=COUNTA(C{start_row}:C{end_row})'
            elif i == 1: val_c.value = f'=COUNTIF({ach_col_letter}{start_row}:{ach_col_letter}{end_row},">0")'
            elif i == 2: 
                val_c.value, val_c.number_format = f'=IFERROR(AVERAGE({ach_col_letter}{start_row}:{ach_col_letter}{end_row}),0)', '0.00%'
            elif i == 3: val_c.value = f'=COUNTIF({col_e_start}{start_row}:{col_e_end}{end_row},"<>" )'
            elif i == 4: 
                val_c.value, val_c.number_format = f'=IFERROR(C{r-1}/(C{r-4}*{num_days}),0)', '0.00%'
            
            if month_num > 1:
                prev_s = f"{year}年{month_num-1}月打卡"
                sum_mom_c.value = f'=IFERROR((C{r}-\'{prev_s}\'!C{r})/\'{prev_s}\'!C{r},0)'

        # 设置列宽
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 30 # 增加事项列宽
        sheet.column_dimensions['D'].width = 12
        for d in range(5, num_days + 5): sheet.column_dimensions[get_column_letter(d)].width = 8 # 增加打卡列宽

    if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])
    workbook.save(filename)
    print(f"Excel模板 '{filename}' 已生成！样式已微调，看起来更宽敞。")

if __name__ == "__main__":
    generate_365_excel_template()
